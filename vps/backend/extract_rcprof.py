#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Estrazione tariffe RC Professionali da Excel -> JSON strutturato."""
import openpyxl, json, re, datetime
from openpyxl.utils import get_column_letter

SRC = '/root/.claude/uploads/808bfd02-4edb-57e0-91b4-f9746907d694/da0783ef-ALL_NEW_TABLE_RATE_ACCREDITED_1.xlsx'
OUT = '/home/user/QUOTE/rcprof_tariffe.json'

wb = openpyxl.load_workbook(SRC, data_only=True)
wbf = openpyxl.load_workbook(SRC, data_only=False)

anomalies = {}  # foglio -> [str]
def warn(sheet, msg):
    anomalies.setdefault(sheet, []).append(msg)

def is_orange(ws_f, r, c):
    cell = ws_f.cell(r, c)
    try:
        return cell.fill.fgColor.rgb == 'FFFFC000'
    except Exception:
        return False

def cell(ws, r, c):
    return ws.cell(r, c).value

def to_num(v):
    """Convert a turnover/maximal header cell to an int when possible."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    s = s.replace('€', '').replace('\xa0', ' ').strip()
    # patterns like '350.000,00' or '250.000 €' or '250000'
    m = re.match(r'^[\d\.\s]+(,\d+)?$', s)
    if m:
        t = s.replace('.', '').replace(' ', '').replace(',', '.')
        try:
            f = float(t)
            return int(f) if f == int(f) else f
        except ValueError:
            return s
    return s

def premium(v):
    """net premium cell -> float or None for // / blanks."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('//', '/', '', '-'):
        return None
    try:
        return float(s.replace('.', '').replace(',', '.'))
    except ValueError:
        return s  # keep raw

def sheet_has_plus10(ws):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and 'Prodotto con 10% Ap sul CRM' in c.value:
                return True
    return False

result = {'fogli': {}, 'note': {}}

# ---------------------------------------------------------------------------
# MODELLO (A): FATTURATO x MASSIMALE
# Generic extractor: scan for "TURNOVER"/"MONTHS" header rows.
# ---------------------------------------------------------------------------

def parse_turnover_block(ws, ws_f, header_row, name, prof_name_cell=None):
    """Parse one turnover x maximal table starting at header_row in left column.
    Returns dict for one profession. header_row is the row with TURNOVER label."""
    sheet = ws.title
    # find label column (the col containing TURNOVER/MONTHS)
    label_col = None
    for c in range(1, ws.max_column + 1):
        v = cell(ws, header_row, c)
        if isinstance(v, str) and v.strip().upper() in ('TURNOVER', 'MONTHS'):
            label_col = c
            break
    if label_col is None:
        return None
    # massimali = consecutive numeric/parseable headers to the right
    massimali = []
    mass_cols = []
    c = label_col + 1
    while c <= ws.max_column:
        v = cell(ws, header_row, c)
        if v is None:
            # allow a single gap? stop on blank
            break
        num = to_num(v)
        massimali.append(num)
        mass_cols.append(c)
        c += 1
    # data rows: below header until DEDUCTIBLE / blank first cell / new section
    tabella = {}
    deductible = {}
    r = header_row + 1
    last_data_row = header_row
    while r <= ws.max_row:
        first = cell(ws, header_row + (r - header_row), label_col)
        first = cell(ws, r, label_col)
        if first is None:
            # could be a gap inside; peek next non-empty
            # stop if two blanks or next is a new header
            nxt = cell(ws, r + 1, label_col)
            if nxt is None:
                break
            r += 1
            continue
        if isinstance(first, str):
            fu = first.strip().upper()
            if fu.startswith('DEDUCTIBLE'):
                for mc, m in zip(mass_cols, massimali):
                    deductible[str(m)] = cell(ws, r, mc)
                last_data_row = r
                r += 1
                continue
            if fu in ('TURNOVER', 'MONTHS') or fu.startswith('MISCELLANEOUS'):
                break
            # other text => end of table
            break
        # numeric turnover row
        key = to_num(first)
        row_vals = {}
        for mc, m in zip(mass_cols, massimali):
            row_vals[str(m)] = premium(cell(ws, r, mc))
        tabella[str(key)] = row_vals
        last_data_row = r
        r += 1

    # accessorie: scan area to the right of the maximals, within row range of table.
    # Stop the horizontal scan at the column where a sibling TURNOVER table begins
    # (e.g. AVVOCATI has a second table starting in col J) and skip the
    # GENERAL SUBLIMIT free-text block.
    accessorie = {}
    start_acc = (mass_cols[-1] + 1) if mass_cols else (label_col + 1)
    r0, r1 = header_row, last_data_row + 1
    # find horizontal limit: a sibling TURNOVER/MONTHS or GENERAL SUBLIMIT header to the right
    acc_limit = ws.max_column
    for cc in range(start_acc, ws.max_column + 1):
        for rr in range(max(1, header_row - 3), r1 + 1):
            v = cell(ws, rr, cc)
            if isinstance(v, str) and v.strip().upper() in ('TURNOVER', 'MONTHS', 'GENERAL SUBLIMIT'):
                acc_limit = cc - 1
                break
        if acc_limit != ws.max_column:
            break
    ACC_KEYS = ('RCO', 'M&A', 'HIGH RISK', 'ACCIDENTAL POLLUTION', 'RETROACT', 'VISTO',
                'HEALTH & SAFETY', 'RECUPERO', 'COVERHOLDER', 'DISCOVERY')
    for rr in range(r0, r1 + 1):
        for cc in range(start_acc, acc_limit + 1):
            v = cell(ws, rr, cc)
            if isinstance(v, str) and v.strip():
                key = v.strip()
                val = cell(ws, rr, cc + 1)
                ku = key.upper()
                if ku == 'GENERAL SUBLIMIT':
                    continue
                # value to the right => accessory name/value pair
                if val is not None and (isinstance(val, (int, float)) or (isinstance(val, str) and val.strip())):
                    accessorie[key] = val if not isinstance(val, datetime.datetime) else str(val)
                else:
                    # continuation line (e.g. 'illimitata +20%' / 'AP 20/30%' under Retroactivity)
                    accessorie[key] = key
    # Also capture accessorie placed BELOW the table in the label column
    # (AVVOCATI: 'Retroactivity'/'RCO' rows beneath, with value in next col).
    for rr in range(last_data_row + 1, min(ws.max_row, last_data_row + 8) + 1):
        v = cell(ws, rr, label_col)
        if isinstance(v, str) and v.strip():
            kv = v.strip()
            ku = kv.upper()
            # skip section headers like 'VISTO LEGGERO*' (asterisk) or generic sublimit titles
            if kv.endswith('*') or ku == 'GENERAL SUBLIMIT':
                continue
            if any(k in ku for k in ACC_KEYS):
                accessorie[kv] = cell(ws, rr, label_col + 1)
            elif ku in ('TURNOVER', 'MONTHS'):
                break
    # plus10: orange anywhere in this block (incl. title rows above) OR text present
    plus10 = False
    title_lo = max(1, header_row - 3)
    for rr in range(title_lo, r1 + 1):
        for cc in range(1, ws.max_column + 1):
            if is_orange(ws_f, rr, cc):
                plus10 = True
            v = cell(ws, rr, cc)
            if isinstance(v, str) and 'Prodotto con 10% Ap sul CRM' in v:
                plus10 = True
    return {
        'nome': name,
        'plus10': plus10,
        'massimali': massimali,
        'tabella': tabella,
        'deductible': deductible,
        'accessorie': accessorie,
    }

def find_profession_name(ws, header_row, label_col=2):
    """Look upward from header_row for the nearest non-empty title in label_col."""
    for r in range(header_row - 1, max(0, header_row - 4), -1):
        v = cell(ws, r, label_col)
        if isinstance(v, str) and v.strip() and v.strip().upper() not in ('TURNOVER', 'MONTHS'):
            if 'PREMI NETTI' in v.upper():
                continue
            return v.strip()
    return None

def extract_turnover_sheet(name, label_cols=(2,)):
    """Generic for TECNICI/AVVOCATI/A.FISCALE/VARIE/SPECIAL MISCELLANEOUS.
    label_cols: which columns can contain TURNOVER (B=2 left table, J=10 right)."""
    ws = wb[name]
    ws_f = wbf[name]
    professioni = []
    notes = []
    for r in range(1, ws.max_row + 1):
        for lc in label_cols:
            v = cell(ws, r, lc)
            if isinstance(v, str) and v.strip().upper() in ('TURNOVER', 'MONTHS'):
                pname = find_profession_name(ws, r, lc)
                blk = parse_turnover_block(ws, ws_f, r, pname or f'{name}_{get_column_letter(lc)}{r}')
                if blk:
                    professioni.append(blk)
    return {'modello': 'turnover', 'professioni': professioni, 'note': notes}

# ---------------------------------------------------------------------------
result['fogli']['TECNICI'] = extract_turnover_sheet('TECNICI', label_cols=(2,))
result['fogli']['AVVOCATI'] = extract_turnover_sheet('AVVOCATI', label_cols=(2, 10))
result['fogli']['A.FISCALE'] = extract_turnover_sheet('A.FISCALE', label_cols=(2, 10))
result['fogli']['VARIE'] = extract_turnover_sheet('VARIE', label_cols=(2,))
result['fogli']['SPECIAL MISCELLANEOUS'] = extract_turnover_sheet('SPECIAL MISCELLANEOUS', label_cols=(2,))

# ---------------------------------------------------------------------------
# A.FISCALE extra special block VISTO LEGGERO (rows ~111-116) - capture raw
# ---------------------------------------------------------------------------
af = wb['A.FISCALE']
visto = {}
for r in range(111, 117):
    rowd = {}
    for c in range(2, 13):
        v = cell(af, r, c)
        if v is not None:
            rowd[get_column_letter(c)] = v if not isinstance(v, datetime.datetime) else str(v)
    if rowd:
        visto[str(r)] = rowd
result['fogli']['A.FISCALE']['blocco_visto_leggero_raw'] = visto
warn('A.FISCALE', "La voce 'VISTO LEGGERO*' (righe 111-116) e' una mini-tabella speciale CON 730 / SENZA 730 (massimale 3.000.000), non una tariffa fatturato x massimale standard: catturata sia come 'professione' degenere sia in 'blocco_visto_leggero_raw'.")
warn('AVVOCATI', "Retroactivity=Unlimited e RCO=0.1 (B46/B47) sono globali per il foglio: associati alle tabelle STUDIO ASSOCIATO; valgono anche per AVVOCATI/AVVOCATO SINDACO.")
warn('AVVOCATI', "Le tabelle STUDIO ASSOCIATO usano massimali testuali '1.000.000/2.000.000' e '2.000.000/2.000.000' (limite singolo/aggregato), conservati come stringhe.")
warn('MEDICI', "Classe G compare in due dimensioni (SENZA ATTI INVASIVI col G e CON ATTI INVASIVI col I): la seconda e' salvata come chiave 'G#CON'. Il file MEDICI non contiene la dimensione 'CON CHIRURGIA'.")
warn('VARIE', "Blocchi non-fatturato (INSURANCE AGENTS/BROKER, CREDIT BROKER, HAULIERS) conservati raw in 'blocchi_speciali_raw'; alcuni usano RATES/PREMIUM MINIMUM invece di premi per massimale.")

# ---------------------------------------------------------------------------
# VARIE: special sub-tables (non turnover x maximal) captured raw
# ---------------------------------------------------------------------------
varie = wb['VARIE']
varie_special = {}
def grab_raw(ws, r0, r1, c0=2, c1=None):
    c1 = c1 or ws.max_column
    block = []
    for r in range(r0, r1 + 1):
        row = {}
        for c in range(c0, c1 + 1):
            v = cell(ws, r, c)
            if v is not None:
                row[get_column_letter(c)] = v if not isinstance(v, datetime.datetime) else str(v)
        if row:
            block.append({'row': r, 'cells': row})
    return block
# Insurance agents/broker (105-113), Credit broker (117-128), Hauliers (131-149),
# DPO (152-159), D&O (162-166)
varie_special['INSURANCE_AGENTS_BROKER'] = grab_raw(varie, 105, 113)
varie_special['CREDIT_BROKER_FINANCIAL_AGENTS'] = grab_raw(varie, 117, 128)
varie_special['HAULIERS'] = grab_raw(varie, 131, 149)
varie_special['DPO'] = grab_raw(varie, 152, 159)
varie_special['D&O'] = grab_raw(varie, 162, 166)
result['fogli']['VARIE']['blocchi_speciali_raw'] = varie_special

# ---------------------------------------------------------------------------
# MISCELLANEO: 3 turnover tables (CAT A/B/C) + category profession lists + new approved
# ---------------------------------------------------------------------------
def extract_miscellaneo():
    name = 'MISCELLANEO'
    ws = wb[name]; ws_f = wbf[name]
    professioni = []
    # the turnover tables use label col A(1), maximals in B..F
    for r in range(1, ws.max_row + 1):
        v = cell(ws, r, 1)
        if isinstance(v, str) and v.strip().upper() == 'TURNOVER':
            # title = nearest 'MISCELLANEOUS (CATEGORY X)' above
            title = None
            for rr in range(r - 1, max(0, r - 4), -1):
                tv = cell(ws, rr, 1)
                if isinstance(tv, str) and 'MISCELLANEOUS' in tv.upper():
                    title = tv.strip(); break
            blk = parse_turnover_block_colA(ws, ws_f, r, title or f'MISC_{r}')
            if blk:
                professioni.append(blk)
    # category lists (cols J/K = list1, M/N = list2). Headers 'CATEGORY A' etc in J/M.
    categorie = {}
    cur_jk = None; cur_mn = None
    for r in range(1, ws.max_row + 1):
        j = cell(ws, r, 10); k = cell(ws, r, 11)
        m = cell(ws, r, 13); n = cell(ws, r, 14)
        # detect category header in J or M
        for colpair, valh in ((('J','K'), j), (('M','N'), m)):
            if isinstance(valh, str) and re.match(r'(?i)^CATEGOR(Y|IA)\s+[A-Z]$', valh.strip()):
                catname = valh.strip().upper().split()[-1]
                categorie.setdefault('CATEGORY_' + catname, [])
                if colpair[0] == 'J':
                    cur_jk = 'CATEGORY_' + catname
                else:
                    cur_mn = 'CATEGORY_' + catname
        # data: english in J/M, italian in K/N
        if cur_jk and isinstance(j, str) and not re.match(r'(?i)^CATEGOR', j.strip()):
            categorie[cur_jk].append({'en': j.strip(), 'it': (k.strip() if isinstance(k, str) else None)})
        if cur_mn and isinstance(m, str) and not re.match(r'(?i)^CATEGOR', m.strip()):
            categorie[cur_mn].append({'en': m.strip(), 'it': (n.strip() if isinstance(n, str) else None)})
    # new approved professions (A78+)
    nuove = []
    for r in range(80, ws.max_row + 1):
        a = cell(ws, r, 1)
        if isinstance(a, str) and a.strip():
            nuove.append({
                'tipologia': a.strip(),
                'categoria': cell(ws, r, 2),
                'assimilazione': cell(ws, r, 3),
                'approvazione': str(cell(ws, r, 4)) if cell(ws, r, 4) is not None else None,
                'assicurato': cell(ws, r, 5),
                'note': cell(ws, r, 6),
            })
    return {'modello': 'turnover', 'professioni': professioni,
            'categorie_professioni': categorie, 'nuove_professioni_approvate': nuove}

def parse_turnover_block_colA(ws, ws_f, header_row, name):
    label_col = 1
    massimali = []; mass_cols = []
    c = label_col + 1
    while c <= ws.max_column:
        v = cell(ws, header_row, c)
        if v is None or (isinstance(v, str) and not re.match(r'^[\d\.\s,€]+$', v.strip())):
            break
        massimali.append(to_num(v)); mass_cols.append(c); c += 1
    tabella = {}; deductible = {}
    r = header_row + 1; last = header_row
    while r <= ws.max_row:
        first = cell(ws, r, label_col)
        if first is None:
            break
        if isinstance(first, str):
            fu = first.strip().upper()
            if fu.startswith('DEDUCTIBLE'):
                for mc, m in zip(mass_cols, massimali):
                    deductible[str(m)] = cell(ws, r, mc)
                last = r; break
            break
        tabella[str(to_num(first))] = {str(m): premium(cell(ws, r, mc)) for mc, m in zip(mass_cols, massimali)}
        last = r; r += 1
    # accessorie in G/H (Retroattività, RCO) within rows
    accessorie = {}
    for rr in range(header_row, last + 1):
        g = cell(ws, rr, 7); h = cell(ws, rr, 8)
        if isinstance(g, str) and g.strip():
            accessorie[g.strip()] = h
        elif h is not None and isinstance(h, str) and h.strip():
            accessorie[h.strip()] = h.strip()
    plus10 = False
    for rr in range(max(1, header_row - 3), last + 1):
        for cc in range(1, 9):
            if is_orange(ws_f, rr, cc):
                plus10 = True
            v = cell(ws, rr, cc)
            if isinstance(v, str) and 'Prodotto con 10% Ap sul CRM' in v:
                plus10 = True
    return {'nome': name, 'plus10': plus10, 'massimali': massimali,
            'tabella': tabella, 'deductible': deductible, 'accessorie': accessorie}

result['fogli']['MISCELLANEO'] = extract_miscellaneo()

# ---------------------------------------------------------------------------
# MEDICI (modello classi)
# ---------------------------------------------------------------------------
def extract_medici():
    ws = wb['MEDICI']
    # global params
    info = {
        'premi': 'lordi',
        'massimale_per_sinistro': cell(ws, 4, 2),
        'massimale_aggregato_annuo': cell(ws, 5, 2),
        'franchigia': cell(ws, 7, 2),
        'off_label_franchigia': cell(ws, 8, 2),
        'retroattivita': {'10 anni': 0, 'illimitata': 0.25},
        'postuma': {'10 anni': 0},
        'prescrizione_off_label': {'tipo': cell(ws, 13, 1), 'AP': cell(ws, 13, 3)},
        'neolaureato': {'premio': 350.0, 'franchigia': 0},
        'specializzando': {'premio': 300.0, 'franchigia': 0},
    }
    # dimension columns: E (5) = SENZA ATTI INVASIVI E SENZA CHIRURGIA,
    # G (7) = CON ATTI INVASIVI E SENZA CHIRURGIA (header label spans E2 / I2),
    # I (9) = CON CHIRURGIA (per header I2 'CON ATTI INVASIVI E SENZA CHIRURGIA' -- see note)
    # Header row 2: E2='SENZA ATTI INVASIVI E SENZA CHIRUGIA', I2='CON ATTI INVASIVI E SENZA CHIRURGIA'
    dim_map = {5: cell(ws, 2, 5), 9: cell(ws, 2, 9)}
    # Columns E,G belong to dim at col5; I,(no further) belong to dim at col9.
    # We'll assign each class column to a dimension by position.
    classi = {}
    class_re = re.compile(r'^([A-Z])\s*-\s*€?\s*([\d\.\s]+,\d{2}|\d[\d\.]*)')
    # scan columns E,G,I for class headers and their specializations beneath
    col_dim = {5: 'SENZA ATTI INVASIVI E SENZA CHIRURGIA',
               7: 'SENZA ATTI INVASIVI E SENZA CHIRURGIA',
               9: 'CON ATTI INVASIVI E SENZA CHIRURGIA'}
    for col in (5, 7, 9):
        cur_class = None
        for r in range(2, ws.max_row + 1):
            v = cell(ws, r, col)
            if not isinstance(v, str) or not v.strip():
                continue
            m = class_re.match(v.strip())
            if m:
                lettera = m.group(1)
                premio = premium(m.group(2))
                cur_class = lettera
                key = lettera
                if key in classi:
                    # same letter may appear in multiple dims; keep first, note dim
                    key = lettera + '#' + col_dim[col][:4]
                classi[key] = {'classe': lettera, 'premio_lordo': premio,
                               'categoria_invasivita': col_dim[col],
                               'specializzazioni': []}
                cur_key = key
            elif cur_class is not None:
                classi[cur_key]['specializzazioni'].append(v.strip())
    return {'modello': 'classi', 'info': info, 'classi': classi}

result['fogli']['MEDICI'] = extract_medici()

# ---------------------------------------------------------------------------
# PARAMEDICI (modello classi)
# ---------------------------------------------------------------------------
def extract_paramedici():
    ws = wb['PARAMEDICI']
    info = {
        'premi': 'lordi',
        'massimale_per_sinistro': cell(ws, 5, 2),
        'massimale_per_anno': cell(ws, 6, 2),
        'retroattivita': {'10 anni gratuita': 0, 'illimitata': 0.25},
        'qualifiche_direttive': {
            'Direttore di Struttura Semplice/Complessa': 50.0,
            'Direttore professioni sanitarie': 50.0,
        },
        'condizioni_operanti': [cell(ws, r, 2) for r in range(19, 24) if cell(ws, r, 2)],
    }
    classi = {}
    class_re = re.compile(r'^([A-Z])\s*-\s*€?\s*([\d\.\s]+,?\d*)')
    # class headers appear in cols F(6),H(8),J(10),L(12) on rows 4,35,46
    cur = {}
    for col in (6, 8, 10, 12):
        cur_key = None
        for r in range(2, ws.max_row + 1):
            v = cell(ws, r, col)
            if not isinstance(v, str) or not v.strip():
                continue
            m = class_re.match(v.strip())
            if m and v.strip().upper() != 'CLASSI DI RISCHIO':
                lettera = m.group(1)
                premio = premium(m.group(2))
                key = lettera
                while key in classi:
                    key = key + "'"
                classi[key] = {'classe': lettera, 'premio_lordo': premio, 'specializzazioni': []}
                cur_key = key
            elif cur_key is not None and v.strip().upper() != 'CLASSI DI RISCHIO':
                classi[cur_key]['specializzazioni'].append(v.strip())
    return {'modello': 'classi', 'info': info, 'classi': classi}

result['fogli']['PARAMEDICI'] = extract_paramedici()

# ---------------------------------------------------------------------------
# CERT 5.0 (speciale)
# ---------------------------------------------------------------------------
def extract_cert50():
    ws = wb['CERT 5.0']
    bande = []
    for r in range(5, 17):
        b = cell(ws, r, 2)
        if b is None:
            continue
        bande.append({
            'credito_imposta_da': cell(ws, r, 2),
            'credito_imposta_a': cell(ws, r, 3),
            'massimale': cell(ws, r, 4),
            'premium_lordo': cell(ws, r, 5),
            'franchigia': cell(ws, r, 6),
        })
    return {'modello': 'speciale', 'premi': 'lordi', 'tipo': "credito d'imposta x massimale",
            'note': cell(ws, 18, 2), 'bande': bande}

result['fogli']['CERT 5.0'] = extract_cert50()

# ---------------------------------------------------------------------------
# RC PATRIMONIALE COLPA GRAVE (speciale)
# ---------------------------------------------------------------------------
def extract_rcpatr():
    ws = wb['RC PATRIMONIALE COLPA GRAVE']
    # premi lordi cols B..G (2..7), premi netti cols I..N (9..14)
    cat_hdr = re.compile(r'(?i)^CATEGORIA\s+([A-Z])$')
    categorie = {}
    cur_cat = None
    massimali_lordi = []; massimali_netti = []
    for r in range(3, ws.max_row + 1):
        a = cell(ws, r, 1)
        if isinstance(a, str) and cat_hdr.match(a.strip()):
            cur_cat = cat_hdr.match(a.strip()).group(1)
            massimali_lordi = [cell(ws, r, c) for c in range(2, 8)]
            massimali_netti = [cell(ws, r, c) for c in range(9, 15)]
            categorie[cur_cat] = {'massimali': [m for m in massimali_lordi], 'ruoli': {}}
            continue
        if isinstance(a, str) and a.strip().upper().startswith(('EXCLUSIONS', 'EXTRA GUARANTEES', 'RETROATTIVITA', 'POSTUMA', 'MORE POSITIONS')):
            break
        if cur_cat and isinstance(a, str) and a.strip():
            ruolo = a.strip()
            mp = {}
            for idx, c in enumerate(range(2, 8)):
                lordo = cell(ws, r, c)
                netto = cell(ws, r, 9 + idx)
                m = massimali_lordi[idx]
                mp[str(m)] = {'lordo': lordo, 'netto': netto}
            categorie[cur_cat]['ruoli'][ruolo] = mp
    extra = {
        'retroattivita': cell(ws, 72, 2),
        'postuma': cell(ws, 73, 2),
        'esclusioni': [cell(ws, r, 1) for r in range(76, 81) if cell(ws, r, 1)],
        'garanzie_aggiuntive': cell(ws, 83, 1),
        'more_positions': cell(ws, 84, 1),
    }
    return {'modello': 'speciale', 'premi': 'lordi e netti', 'categorie': categorie, 'extra': extra}

result['fogli']['RC PATRIMONIALE COLPA GRAVE'] = extract_rcpatr()

# ---------------------------------------------------------------------------
# D&O (speciale)
# ---------------------------------------------------------------------------
def extract_do():
    ws = wb['D&O']
    # suggested tariffs: header row 3 cols B..H, rows 4..7
    massimali = [cell(ws, 3, c) for c in range(2, 9)]
    tabella = {}
    for r in range(4, 8):
        ta = cell(ws, r, 1)
        if ta is None:
            continue
        tabella[str(ta).strip()] = {str(m): cell(ws, r, c) for m, c in zip(massimali, range(2, 9))}
    # original AIEL
    massimali_orig = [cell(ws, 12, c) for c in range(2, 7)]
    orig = {}
    ta = cell(ws, 13, 1)
    if ta:
        orig[str(ta).strip()] = {str(m): cell(ws, 13, c) for m, c in zip(massimali_orig, range(2, 7))}
    return {'modello': 'speciale', 'tipo': 'total assets x massimale',
            'suggested': {'massimali': massimali, 'tabella': tabella},
            'original_AIEL': {'massimali': massimali_orig, 'tabella': orig},
            'retro': {'Retro 0': 0, 'Retro illimitata': 0.20}}

result['fogli']['D&O'] = extract_do()

# ---------------------------------------------------------------------------
# Lettere E (speciale)
# ---------------------------------------------------------------------------
def extract_lettereE():
    ws = wb['Lettere E']
    righe = []
    for r in range(5, 12):
        b = cell(ws, r, 2)
        if b is None:
            continue
        righe.append({
            'fascia_massimale': b,
            'premio_C': cell(ws, r, 3),
            'premio_D_1M': cell(ws, r, 4),
            'coef_E': cell(ws, r, 5),
            'coef_F': cell(ws, r, 6),
            'premio_G_2_5M': cell(ws, r, 7),
        })
    return {'modello': 'speciale', 'premi': 'lordi',
            'colonne_massimali': {'D': cell(ws, 4, 4), 'G': cell(ws, 4, 7)},
            'franchigia': 1000, 'righe': righe}

result['fogli']['Lettere E'] = extract_lettereE()

# ---------------------------------------------------------------------------
# CONSISTENT RATES (speciale - motore alternativo)
# ---------------------------------------------------------------------------
def extract_consistent():
    ws = wb['CONSISTENT RATES']
    # fattori: B5.. policy limit, C/D/E LOW/MEDIUM/HIGH
    fattori = []
    for r in range(5, 20):
        pl = cell(ws, r, 2)
        if pl is None:
            continue
        fattori.append({'policy_limit': pl, 'LOW': cell(ws, r, 3),
                        'MEDIUM': cell(ws, r, 4), 'HIGH': cell(ws, r, 5)})
    # fattori base header
    fattori_base = {'LOW': cell(ws, 3, 3), 'MEDIUM': cell(ws, 3, 4), 'HIGH': cell(ws, 3, 5)}
    # sconti incrementali: G/H min/max, I incremental discount, J incr rate, K cum premium, L cum rate
    sconti = []
    for r in range(4, ws.max_row + 1):
        mn = cell(ws, r, 7)
        if mn is None:
            continue
        sconti.append({
            'min': cell(ws, r, 7), 'max': cell(ws, r, 8),
            'incremental_discount': cell(ws, r, 9),
            'incremental_rate': cell(ws, r, 10),
            'cumulative_premium': cell(ws, r, 11),
            'cumulative_rate': cell(ws, r, 12),
        })
    return {'modello': 'speciale', 'tipo': 'motore alternativo',
            'massimale_base': 1000000,
            'fattori_massimale': {'base_factor': fattori_base, 'tabella': fattori},
            'sconti_incrementali_fatturato': sconti}

result['fogli']['CONSISTENT RATES'] = extract_consistent()

# ---------------------------------------------------------------------------
# NOTE
# ---------------------------------------------------------------------------
result['note'] = {
    'fonte': SRC,
    'data_estrazione': str(datetime.date.today()),
    'plus10': "true se il foglio/tabella contiene 'Prodotto con 10% Ap sul CRM' o ha celle arancioni FFC000 (=+10% Ap sul CRM finale)",
    'celle_non_disponibili': "// / o vuote => null nei premi netti",
    'accessorie_codici': {
        'RCO 0.1': '+10% sul netto',
        'M&A 0.2': '+20%',
        'High Risks 0.25': '+25%',
        'Accidental Pollution 0.15': '+15%',
        'Visto Leggero/Pesante': 'euro fissi sul netto (50 / 150)',
        'Retroactivity illimitata +20%': 'maggiorazione retroattivita',
    },
    'anomalie': anomalies,
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print('JSON salvato in', OUT)
